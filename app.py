import nmap
import json
import os
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

init()  # Initialize colorama

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///scans.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
os.makedirs('reports', exist_ok=True)

# Custom template filter for basename
@app.template_filter('basename')
def basename_filter(path):
    return os.path.basename(path)

# Database model for scan history
class Scan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    target = db.Column(db.String(255))
    ports = db.Column(db.String(255))
    verbose = db.Column(db.Boolean)
    os_detection = db.Column(db.Boolean)
    service_detection = db.Column(db.Boolean)
    vuln_scan = db.Column(db.Boolean)
    results = db.Column(db.Text)  # Store results as JSON string

# Create database tables
with app.app_context():
    db.create_all()

def scan_target(target, ports='1-1024', verbose=False, os_detection=False, service_detection=False, vuln_scan=False):
    """Performs a security scan using Nmap."""
    scanner = nmap.PortScanner()
    args = f'-p {ports}'
    if verbose:
        args += ' -v'
    if os_detection:
        args += ' -O'
    if service_detection:
        args += ' -sV'
    if vuln_scan:
        args += ' --script vuln'
    
    print(f"{Fore.GREEN}Starting scan on {target} with args: {args}{Style.RESET_ALL}")
    try:
        scanner.scan(target, arguments=args)
    except nmap.PortScannerError as e:
        return {'error': str(e)}
    
    results = {}
    for host in scanner.all_hosts():
        results[host] = {
            'status': scanner[host].state(),
            'hostname': scanner[host].hostname(),
            'protocols': {}
        }
        if os_detection and 'osclass' in scanner[host]:
            results[host]['os'] = scanner[host]['osclass']
        
        for proto in scanner[host].all_protocols():
            results[host]['protocols'][proto] = {}
            ports = scanner[host][proto].keys()
            for port in ports:
                port_info = scanner[host][proto][port]
                results[host]['protocols'][proto][port] = {
                    'state': port_info['state'],
                    'name': port_info['name'],
                    'product': port_info.get('product', 'unknown'),
                    'version': port_info.get('version', 'unknown')
                }
                if vuln_scan and 'script' in port_info:
                    results[host]['protocols'][proto][port]['vulnerabilities'] = port_info['script']
    
    return results

def generate_excel_report(scan):
    """Generate an Excel report for a given Scan object."""
    results = json.loads(scan.results)
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Report"

    # Define headers
    headers = ["Host", "Hostname", "Status", "Protocol", "Port", "Service", "Product", "Version", "State", "OS Info", "Vulnerabilities"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    # Add scan metadata
    ws.append(["Scan Metadata", "", "", "", "", "", "", "", "", "", ""])
    ws.append(["Target", scan.target, "", "", "", "", "", "", "", "", ""])
    ws.append(["Timestamp", str(scan.timestamp), "", "", "", "", "", "", "", "", ""])
    ws.append(["Ports", scan.ports, "", "", "", "", "", "", "", "", ""])
    ws.append(["Options", f"Verbose: {scan.verbose}, OS: {scan.os_detection}, Service: {scan.service_detection}, Vuln: {scan.vuln_scan}", "", "", "", "", "", "", "", "", ""])
    ws.append([])  # Empty row for spacing

    # Add scan results
    row = ws.max_row + 1
    for host, host_info in results.items():
        status = host_info.get('status', 'unknown')
        hostname = host_info.get('hostname', '')
        os_info = json.dumps(host_info.get('os', {}), indent=2) if 'os' in host_info else ""

        for proto, ports in host_info.get('protocols', {}).items():
            for port, info in ports.items():
                vulnerabilities = json.dumps(info.get('vulnerabilities', {}), indent=2) if 'vulnerabilities' in info else ""
                ws.append([
                    host,
                    hostname,
                    status,
                    proto,
                    port,
                    info.get('name', ''),
                    info.get('product', ''),
                    info.get('version', ''),
                    info.get('state', ''),
                    os_info,
                    vulnerabilities
                ])
                os_info = ""  # Only include OS info for the first row of each host

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50
        ws.column_dimensions[column].width = adjusted_width

    filename = f"scan_report_{scan.id}.xlsx"
    wb.save(f"reports/{filename}")
    return filename

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        target = request.form.get('target')
        ports = request.form.get('ports', '1-1024')
        verbose = 'verbose' in request.form
        os_detection = 'os_detection' in request.form
        service_detection = 'service_detection' in request.form
        vuln_scan = 'vuln_scan' in request.form

        if not target:
            return render_template('index.html', error="Please provide a target IP or hostname")

        results = scan_target(target, ports, verbose, os_detection, service_detection, vuln_scan)
        
        if 'error' in results:
            return render_template('index.html', error=results['error'])

        # Store in database
        new_scan = Scan(
            target=target,
            ports=ports,
            verbose=verbose,
            os_detection=os_detection,
            service_detection=service_detection,
            vuln_scan=vuln_scan,
            results=json.dumps(results)
        )
        db.session.add(new_scan)
        db.session.commit()

        # Generate reports
        excel_file = generate_excel_report(new_scan)

        return render_template('index.html', result=results, excel_file=excel_file)

    return render_template('index.html')

@app.route('/history')
def history():
    scans = Scan.query.order_by(Scan.timestamp.desc()).all()
    return render_template('history.html', scans=scans)

@app.route('/download/<int:scan_id>')
def download_report(scan_id):
    scan = Scan.query.get_or_404(scan_id)
    os.makedirs('reports', exist_ok=True)
    filename = generate_excel_report(scan)
    return send_from_directory('reports', filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=1818, debug=True)